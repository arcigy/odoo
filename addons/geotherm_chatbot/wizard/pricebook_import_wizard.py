import base64
import io
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook
from odoo import fields, models
from odoo.exceptions import UserError

from ..models.pricebook import stable_key


RESERVED_SHEETS = {"navod", "oblasti", "otazky", "sablona-sluzby", "doplnky"}
PRODUCT_HEADERS = {
    "stav",
    "nazov-produktu-balika",
    "znacka-model",
    "cena-s-dph",
    "typ-ceny",
    "vhodne-pre",
}


def normalized(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def is_active(value):
    return normalized(value) in {"ano", "true", "1", "aktivne", "aktivna", "active"}


def as_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:[.,]\d+)?", str(value).replace(" ", ""))
    return float(match.group(0).replace(",", ".")) if match else None


def as_date(value):
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if value:
        try:
            return fields.Date.to_date(value).isoformat()
        except (TypeError, ValueError):
            return fields.Date.today().isoformat()
    return fields.Date.today().isoformat()


def source_from_note(value):
    note = str(value or "").strip()
    match = re.search(r"https?://\S+", note)
    source_url = match.group(0).rstrip(".,;)\"") if match else ""
    clean_note = re.sub(r"\s*Zdroj existencie:\s*https?://\S+", "", note, flags=re.I).strip()
    return clean_note, source_url


def infer_series(name, brand_model):
    text = normalized(f"{name} {brand_model}")
    if "arotherm-split-plus" in text:
        return "split_plus"
    if "arotherm-plus" in text:
        return "plus"
    if "arotherm-pro" in text:
        return "pro"
    return stable_key(name, brand_model)


def rows_as_dicts(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalized(value) for value in rows[0]]
    return [{headers[index]: value for index, value in enumerate(row) if index < len(headers)} for row in rows[1:]]


class GeothermPricebookImportWizard(models.TransientModel):
    _name = "geotherm.pricebook.import.wizard"
    _description = "Import Geotherm pricebook workbook"

    file_data = fields.Binary(string="Excel cenník", required=True, attachment=False)
    file_name = fields.Char(required=True)
    mode = fields.Selection(
        [("upsert", "Doplniť a aktualizovať"), ("replace", "Nahradiť celý cenník")],
        default="upsert",
        required=True,
    )

    def action_import(self):
        self.ensure_one()
        if not self.file_name.lower().endswith(".xlsx"):
            raise UserError("Nahrajte súbor vo formáte .xlsx.")
        raw = base64.b64decode(self.file_data)
        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            payload, detected_format = self._parse_workbook(workbook)
            effective_mode = "upsert" if detected_format == "vaillant-original" else self.mode
            counts = self.env["geotherm.pricebook.area"].sudo().import_catalog(
                payload,
                source_file=self.file_name,
                mode=effective_mode,
            )
            attachment = self.env["ir.attachment"].sudo().create({
                "name": self.file_name,
                "datas": self.file_data,
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "res_model": "geotherm.pricebook.import",
            })
            summary = (
                f"Formát: {detected_format}. Oblasti: {counts['areas']}, otázky: {counts['questions']}, "
                f"produkty: {counts['products']}, doplnky: {counts['addons']}."
            )
            log = self.env["geotherm.pricebook.import"].sudo().create({
                "name": self.file_name,
                "mode": effective_mode,
                "state": "done",
                "result_summary": summary,
                "attachment_id": attachment.id,
            })
            attachment.write({"res_id": log.id})
        except Exception as error:
            raise UserError(f"Excel sa nepodarilo importovať: {error}") from error
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {"title": "Cenník bol importovaný", "message": summary, "type": "success", "sticky": False},
        }

    def _parse_workbook(self, workbook):
        normalized_names = {normalized(name): name for name in workbook.sheetnames}
        if "oblasti" in normalized_names and "otazky" in normalized_names:
            return self._parse_business_template(workbook, normalized_names), "geotherm-template"
        if any("arotherm" in normalized(name) for name in workbook.sheetnames):
            return self._parse_original_vaillant(workbook), "vaillant-original"
        raise ValueError("Neznámy Excel. Očakáva sa Geotherm šablóna alebo pôvodný Vaillant cenník.")

    def _parse_business_template(self, workbook, names):
        payload = {
            "version": fields.Datetime.now().isoformat(),
            "source": {"file": self.file_name, "importedAt": fields.Date.today().isoformat(), "notes": ["Importované cez Odoo ERP."]},
            "areas": [],
            "questions": [],
            "products": [],
            "addons": [],
        }
        area_sheet = workbook[names["oblasti"]]
        for row in rows_as_dicts(area_sheet):
            name = str(row.get("nazov-oblasti") or "").strip()
            if name:
                payload["areas"].append({"name": name, "active": is_active(row.get("aktivna")), "note": row.get("poznamka") or ""})

        question_sheet = workbook[names["otazky"]]
        for row in rows_as_dicts(question_sheet):
            area = str(row.get("oblast") or "").strip()
            filter_key = str(row.get("nazov-filtra") or "").strip()
            question = str(row.get("otazka-pre-chatbota") or "").strip()
            if not area or not filter_key or not question:
                continue
            answer_type = str(row.get("typ-odpovede") or "text").strip().lower()
            if answer_type not in {"text", "číslo", "výber", "áno/nie"}:
                answer_type = {"cislo": "číslo", "vyber": "výber", "ano-nie": "áno/nie"}.get(normalized(answer_type), "text")
            payload["questions"].append({
                "active": is_active(row.get("aktivna")),
                "area": area,
                "filter": filter_key,
                "label": str(row.get("co-zistit") or filter_key),
                "question": question,
                "answerType": answer_type,
                "allowedAnswers": [value.strip() for value in str(row.get("mozne-odpovede") or "").split(";") if value.strip()],
                "askWhen": str(row.get("kedy-sa-pytat") or "vždy"),
                "requiredBeforePrice": is_active(row.get("povinne-pred-cenou")),
                "order": int(as_number(row.get("poradie")) or 10),
                "examples": str(row.get("priklady-odpovede") or ""),
            })

        for sheet in workbook.worksheets:
            sheet_key = normalized(sheet.title)
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = {normalized(value) for value in rows[0] if value}
            if sheet_key in RESERVED_SHEETS or not PRODUCT_HEADERS.issubset(headers):
                continue
            area = sheet.title.strip()
            if not any(item["name"] == area for item in payload["areas"]):
                payload["areas"].append({"name": area, "active": True})
            for row in rows_as_dicts(sheet):
                state = normalized(row.get("stav"))
                name = str(row.get("nazov-produktu-balika") or "").strip()
                if not name or state == "vzor":
                    continue
                brand_model = str(row.get("znacka-model") or "").strip()
                filters = {}
                for index in range(1, 4):
                    key = str(row.get(f"vlastny-filter-{index}") or "").strip()
                    value = str(row.get(f"hodnota-filtra-{index}") or "").strip()
                    if key and value:
                        filters[key] = value
                note, source_url = source_from_note(row.get("poznamka-pre-chatbota"))
                payload["products"].append({
                    "active": is_active(row.get("stav")),
                    "area": area,
                    "series": infer_series(name, brand_model),
                    "name": name,
                    "brandModel": brand_model,
                    "priceEurWithVat": as_number(row.get("cena-s-dph")) or 0,
                    "priceType": str(row.get("typ-ceny") or "individuálne"),
                    "fit": str(row.get("vhodne-pre") or ""),
                    "areaFromM2": as_number(row.get("plocha-od-m2")),
                    "areaToM2": as_number(row.get("plocha-do-m2")),
                    "occupantsFrom": as_number(row.get("osob-od")),
                    "occupantsTo": as_number(row.get("osob-do")),
                    "installationIncluded": is_active(row.get("montaz-v-cene")),
                    "included": str(row.get("co-je-v-cene") or ""),
                    "doNotRecommendWhen": str(row.get("kedy-neodporucat") or ""),
                    "chatbotNote": note,
                    "validFrom": as_date(row.get("platne-od")),
                    "customFilters": filters,
                    "priceConfidence": "testovací odhad" if "testovac" in normalized(note) else "klientský cenník",
                    "sourceUrl": source_url,
                })

        addon_name = names.get("doplnky")
        if addon_name:
            for row in rows_as_dicts(workbook[addon_name]):
                area = str(row.get("oblast") or "").strip()
                name = str(row.get("nazov-doplnku") or "").strip()
                if not area or not name:
                    continue
                note, source_url = source_from_note(row.get("poznamka-pre-chatbota"))
                payload["addons"].append({
                    "active": is_active(row.get("stav")),
                    "area": area,
                    "name": name,
                    "priceEurWithVat": as_number(row.get("cena-s-dph")) or 0,
                    "priceType": str(row.get("typ-ceny") or "za ks"),
                    "offerWhen": str(row.get("kedy-ponuknut") or ""),
                    "chatbotNote": note,
                    "priceConfidence": "testovací odhad" if "testovac" in normalized(note) else "klientský cenník",
                    "sourceUrl": source_url,
                })
        return payload

    def _parse_original_vaillant(self, workbook):
        payload = {
            "version": fields.Datetime.now().isoformat(),
            "source": {
                "file": self.file_name,
                "importedAt": fields.Date.today().isoformat(),
                "notes": ["Pôvodný Vaillant Excel. Nejasný 1850 l údaj a posledný aroTHERM pro blok sa neberú ako potvrdený fakt."],
            },
            "areas": [{"name": "Tepelné čerpadlá", "active": True}],
            "questions": [],
            "products": [],
            "addons": [],
        }
        for sheet in workbook.worksheets:
            sheet_key = normalized(sheet.title)
            if "arotherm" not in sheet_key:
                continue
            series = "split_plus" if "split-plus" in sheet_key else "plus" if "plus" in sheet_key else "pro"
            rows = list(sheet.iter_rows(values_only=True))
            index = 0
            while index < len(rows):
                first = str(rows[index][0] or "").strip()
                if not first.lower().startswith("tepelné čerpadlo"):
                    index += 1
                    continue
                price = as_number(rows[index][5] if len(rows[index]) > 5 else None)
                block = []
                cursor = index + 1
                while cursor < len(rows) and cursor <= index + 9 and any(value not in (None, "") for value in rows[cursor]):
                    block.append(str(rows[cursor][0] or "").strip())
                    cursor += 1
                area_text = next((value for value in block if "rozlohe" in normalized(value)), "")
                range_values = re.findall(r"\d+", area_text)
                included = [value.lstrip("- ") for value in block if value.startswith("-")]
                model = next((value for value in reversed(block) if value and not value.startswith("-") and "rozlohe" not in normalized(value)), "")
                ambiguous = series == "pro" and range_values == ["260", "320"]
                if "1850l" in normalized(" ".join(included)):
                    included = [re.sub(r"1850\s*l", "objem zásobníka treba potvrdiť", value, flags=re.I) for value in included]
                payload["products"].append({
                    "active": bool(price) and not ambiguous,
                    "area": "Tepelné čerpadlá",
                    "series": series,
                    "name": first,
                    "brandModel": model,
                    "priceEurWithVat": price or 0,
                    "priceType": "balík",
                    "fit": f"dom približne {range_values[0]} až {range_values[1]} m²" if len(range_values) >= 2 else "rodinný dom",
                    "areaFromM2": float(range_values[0]) if len(range_values) >= 2 else None,
                    "areaToM2": float(range_values[1]) if len(range_values) >= 2 else None,
                    "occupantsFrom": None,
                    "occupantsTo": None,
                    "installationIncluded": True,
                    "included": "; ".join(included),
                    "doNotRecommendWhen": "bez overenia tepelnej straty, vykurovacej sústavy a montážnych podmienok",
                    "chatbotNote": "Nejasný blok, nepoužívať ako istý fakt." if ambiguous else "Orientačná cena s montážou a DPH z klientského cenníka.",
                    "validFrom": fields.Date.today().isoformat(),
                    "customFilters": {},
                    "priceConfidence": "klientský cenník",
                    "sourceUrl": "",
                })
                index = max(cursor, index + 1)
        return payload

